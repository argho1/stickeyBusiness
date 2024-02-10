import os
import sys
import time
import openpyxl
import threading
import pandas as pd
from barcode import Code128
from reportlab.lib import colors
from reportlab.lib.units import mm
from reportlab.pdfgen import canvas
from barcode.writer import ImageWriter
from reportlab.lib.pagesizes import A4
from PyPDF2 import PdfWriter, PdfReader


# Function to print the progress bar and estimate time of completion
def print_progress_bar(page, start_time, total_pages):
    # Current progress
    percentage = (page / total_pages) * 100
    bar_length = 50  # Length of the progress bar
    filled_length = int(bar_length * page // total_pages)
    bar = '█' * filled_length + '-' * (bar_length - filled_length)
    
    # Time calculations
    current_time = time.time()
    elapsed_time = current_time - start_time
    if page > 0:
        estimated_total_time = elapsed_time / page * total_pages
        remaining_time = estimated_total_time - elapsed_time
    else:
        remaining_time = 0  # Just to handle division by zero for the first page
    
    # Formatting elapsed and remaining time
    elapsed_time_formatted = time.strftime("%H:%M:%S", time.gmtime(elapsed_time))
    remaining_time_formatted = time.strftime("%H:%M:%S", time.gmtime(remaining_time))
    
    # Print the progress bar with elapsed time and estimated time of completion
    sys.stdout.write(f'\rPage {page}/{total_pages} |[{bar}| {percentage:.2f}% Completed. Elapsed: {elapsed_time_formatted}, Remaining: {remaining_time_formatted}')
    sys.stdout.flush()


def boxStickers(data, ctnno):

    start_time = time.time()
    
    msn = "M5005491008BKA00" #Add without box number
           
    ean = "0796554198316"

    #ctnno = input("Enter Carton No. : ")

    # Create PDF canvas
    pdf = canvas.Canvas(f'Cartion_Box_Sticker_pdf/Box{ctnno}.pdf', pagesize=A4)




    #collecting data fomr excel file and storing it in an array
    val1 = []
    val2 = []

    val1 = data['SN'][1:]
    val2 = data['MAC'][1:]

    no_of_barcode = len(val1)


    ##code to make 14x per page pdf.
    #change
    barcode_per_page = 12 #int(input("Enter Barcode Per Page (12 max) :"))

    

    if int(ctnno) <= 9:
        ctnno = "0"+str(ctnno)

    msn = (msn+ str(ctnno))

    #no = int(input('Start RSN for this sheet : '))


    
    # Variables to track page count and barcode count
    page_count = 0
    barcode_count = 0

    # Iterate through rows in the Excel sheet
    i = 0

    font_size = 10

    # Constants for positioning barcodes on the page
    x_start = 10 * mm  # Starting x position (20mm from the left margin)
    y_start = 235 * mm  # Starting y position (280mm from the bottom margin)
    x_increment = 0  # No horizontal spacing between barcodes in a row
    y_increment = 18 * mm  # Vertical spacing between barcodes (adjusted to make them closer)
    barcode_width = 90 * mm  # Barcode width
    barcode_height = 16 * mm  # Reduced barcode height

    #wight of one in kg
    oneBox_Gross_Weight = 1.241
    oneBox_Net_Weight = 1.016



    while i < len(val1): 

        # Calculate position on the page
        x = x_start+10 # Adjust space from right
        y = y_start - ((barcode_count % barcode_per_page) * y_increment) - 50  # Adjusted to place 14 barcodes in a column

        x1 = x_start+290 # Adjust space from right
        y1 = y_start - ((barcode_count % barcode_per_page) * y_increment) - 50

        x2 = 20
        y2 = 800
        
        ## Above text left ##
        pdf.setFont("Helvetica", 15)

        pdf.drawString(x2, y2, f"Commodity: Credo CR-3120-OD")
        pdf.drawString(x2, y2-20, f"Color: White")
        pdf.drawString(x2, y2-40, f"PO: I03/450054910")
        pdf.drawString(x2, y2-60, f"Date:02/2024")

        pdf.drawString(x2, y2-100, f"Gross Wt : {round(oneBox_Gross_Weight * no_of_barcode, 2)} Kg")
        pdf.drawString(x2, y2-120, f"Net Wt. : {round(oneBox_Net_Weight * no_of_barcode, 2)} Kg")

        ## Above text right ##
        pdf.drawString(x2+400, y2, f"Carton No. : {ctnno} of 17") #make dynamic value
        pdf.drawString(x2+480, y2-20, f"Qty : {no_of_barcode}")


        msn1 = Code128(msn, writer=ImageWriter())
        msn_image = msn1.render(writer_options={'module_width': 4, 'module_height': 80, "font_size": 20*5, "text_distance": 40, "quite_zone": 10})
        msn_image_filename = (f"./bufferDEL/msn_barcode.png")
        msn_image.save(msn_image_filename)
        pdf.drawImage(msn_image_filename, x2+310, y2-75, width=barcode_width, height=barcode_height)
        pdf.drawString(x2+260, y2-50, f"MSN:")


        ean1 = Code128(ean, writer=ImageWriter())
        ean_image = ean1.render(writer_options={'module_width': 4, 'module_height': 80, "font_size": 20*5, "text_distance": 40, "quite_zone": 10})
        ean_image_filename = (f"./bufferDEL/ean_barcode.png")
        ean_image.save(ean_image_filename)
        pdf.drawImage(ean_image_filename, x2+310, y2-125, width=barcode_width, height=barcode_height)
        pdf.drawString(x2+260, y2-100, f"EAN:")



        pdf.setFont("Helvetica-Bold", font_size-1)
        
        ## RSN placed on the left side of the page. 

        # Insert barcode image into PDF
        rsn = Code128(val1[i], writer=ImageWriter())
        rsn_image = rsn.render(writer_options={'module_width': 2.8, 'module_height': 80, "font_size": 20*5  , "text_distance": 40, "quite_zone": 10})
        rsn_image_filename = (f"./bufferDEL/rsn_barcode_{i}.png")
        rsn_image.save(rsn_image_filename)
        pdf.drawImage(rsn_image_filename, x, y-18, width=barcode_width, height=barcode_height+12)
        pdf.drawString(x-35, y+25, f"RSN:")
        #pdf.drawString(x-30, y+15, f"{no}")


        macid = Code128(val2[i], writer=ImageWriter())
        macid_image = macid.render(writer_options={'module_width': 2.8, 'module_height': 80, "font_size": 20*5, "text_distance": 40, "quite_zone": 10})
        macid_image_filename = (f"./bufferDEL/macid_barcode{i}.png")
        macid_image.save(macid_image_filename)
        pdf.drawImage(macid_image_filename, x1, y1-18, width=barcode_width, height=barcode_height+12)   
        pdf.drawString(x2+311, y2-130, f"MAC ID: ")
        
        







        # Remove the barcode image file
        os.remove(rsn_image_filename)
        os.remove(macid_image_filename)

        # Increment the barcode count
        barcode_count += 1

        # Check if a new page is needed
        if barcode_count == barcode_per_page:
            # Reset parameters for new page
            page_count += 1
            barcode_count = 0

            # Show the current page
            pdf.showPage()




        print_progress_bar(i+1, start_time, no_of_barcode)
        #no = no+1
        i = i + 1 
    # Save the PDF
    print()
    print(f"File save at : Cartion_Box_Sticker_pdf/Box{ctnno}.pdf")
    pdf.save()


# Function to extract data below the search values in the same column
def extract_data_below_values(sheet, search_values):
    # Iterate through all cells and search for the specified values
    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
        for cell in row:
            # Normalize the cell value for comparison
            normalized_cell_value = str(cell.value).lower().replace(' ', '').replace('.', '').replace('\xa0', '')

            # Search for the normalized values
            if any(search_value in normalized_cell_value for search_value in search_values):
                # Get the row and column indices
                row_index = cell.row
                col_index = cell.column

                # Extract the data below the values in the same column
                data = [sheet.cell(i, col_index).value for i in range(row_index + 1, sheet.max_row + 1)]

                # Return a tuple with both data and column index
                return pd.DataFrame({f"{sheet.title}_{col_index}": data}), col_index

    # If none of the search values are found, extract all non-empty values from the first column
    first_column_data = [sheet.cell(i, 1).value for i in range(1, sheet.max_row + 1) if sheet.cell(i, 1).value is not None]

    # Return a DataFrame with the first column data and None for the column index
    return pd.DataFrame({f"{sheet.title}_1": first_column_data}), None


# Function to process the Excel data and extract SN and MAC for each box
def extract_sn_mac(excel_data):
    # Dictionary to hold the box data
    box_data = {}
    boxvalue4search = ['box', 'box no']  # List of strings to search for
    current_box = None  # To keep track of the current box number

    for index, row in excel_data.iterrows():
        # Check if any of the terms in boxvalue4search are in the first cell of the row
        if any(search_term in str(row[0]).lower() for search_term in boxvalue4search):
            current_box = str(row[0])  # Update the current box number
            box_data[current_box] = {'SN': [], 'MAC': []}
        elif pd.notnull(row[1]) and pd.notnull(row[2]) and current_box:
            box_data[current_box]['SN'].append(row[1])
            box_data[current_box]['MAC'].append(row[2])

    return box_data


# Main code integration
startFrom = int(input("Enter Box no. to start with : "))  # Assuming we are starting from box 1 for the sake of demonstration


# Load the entire workbook once, instead of in the loop
location = 'carton_jio300_7feb24.xlsx'
df = pd.read_excel(location)  # You would use the actual path to your Excel file

# Extract the data for all boxes once
extracted_data = extract_sn_mac(df)

# Extract box numbers and convert them to integers for proper numeric sorting
box_numbers = [int(''.join(filter(str.isdigit, box))) for box in extracted_data.keys()]

# Now sort the box numbers in numeric order
sorted_box_numbers = sorted(box_numbers)

for box_number in sorted_box_numbers:
    # Convert back to the original box format if needed, or directly use box_number if applicable
    box_key = f"BOX {box_number}"  # Adjust format as necessary based on how your keys are structured
    if box_number >= startFrom:
        print(f"BOX {box_number}")
        boxStickers(extracted_data[f"BOX {box_number}"], str(box_number))


# Function to print the data in a formatted way
def print_formatted_data(box_data):
    for box, details in box_data.items():
        print(f"Data for {box}:")
        print(f"{'SN':<20} {'MAC':<20}")
        # Skip the first element of each list because it's the header based on how we simulated the data
        for sn, mac in zip(details['SN'][1:], details['MAC'][1:]):
            print(f"{sn:<20} {mac:<20}")
        print("\n")


# Call the function with the extracted data
#print_formatted_data(extracted_data)