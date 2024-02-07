import os
import time
import openpyxl
import pandas as pd
from barcode import Code128
from reportlab.lib import colors
from reportlab.lib.units import mm
from reportlab.pdfgen import canvas
from barcode.writer import ImageWriter
from reportlab.lib.pagesizes import A4
from PyPDF2 import PdfWriter, PdfReader




def boxStickers(sheet, val1, val2):
    msn = input("Enter MSN plz :")

    ean = "0796554198316"


    # Create PDF canvas
    pdf = canvas.Canvas('cartonStickers.pdf', pagesize=A4)

    # Constants for positioning barcodes on the page
    x_start = 10 * mm  # Starting x position (20mm from the left margin)
    y_start = 235 * mm  # Starting y position (280mm from the bottom margin)
    x_increment = 0  # No horizontal spacing between barcodes in a row
    y_increment = 18 * mm  # Vertical spacing between barcodes (adjusted to make them closer)
    barcode_width = 90 * mm  # Barcode width
    barcode_height = 16 * mm  # Reduced barcode height






    ##code to make 14x per page pdf.


    #change
    barcode_per_page = 12 #int(input("Enter Barcode Per Page (12 max) :"))


    if barcode_per_page > 12 or barcode_per_page == '':
        print("Barcode will exceed page limit, setting count to 12!!!")
        barcode_per_page = 12

    # Variables to track page count and barcode count
    page_count = 0
    barcode_count = 0

    ctnno = input("Enter Carton No. : ")
    no = int(input('Start RSN for this sheet : '))

    font_size = 10



    # Iterate through rows in the Excel sheet
    i = 0
    while i < len(val1): 

        # Calculate position on the page
        x = x_start+10 # Adjust space from right
        y = y_start - ((barcode_count % barcode_per_page) * y_increment) - 50  # Adjusted to place 14 barcodes in a column

        x1 = x_start+290 # Adjust space from right
        y1 = y_start - ((barcode_count % barcode_per_page) * y_increment) - 50

        x2 = 20
        y2 = 800
        
        ## Above text left ##
        pdf.setFont("Helvetica", 15)

        pdf.drawString(x2, y2, f"Commodity: Credo CR-3120-OD")
        pdf.drawString(x2, y2-20, f"Color: White")
        pdf.drawString(x2, y2-40, f"PO: I03/450048788")
        pdf.drawString(x2, y2-60, f"Date:01/2024")

        pdf.drawString(x2, y2-100, f"Gross Wt : 22.34 Kg")
        pdf.drawString(x2, y2-120, f"Net Wt. : 20.56 Kg")

        ## Above text right ##
        pdf.drawString(x2+400, y2, f"Carton No. : {ctnno} of 42")
        pdf.drawString(x2+480, y2-20, f"Qty : 12")


        msn1 = Code128(msn, writer=ImageWriter())
        msn_image = msn1.render(writer_options={'module_width': 4, 'module_height': 80, "font_size": 20*5, "text_distance": 40, "quite_zone": 10})
        msn_image_filename = (f"./bufferDEL/msn_barcode.png")
        msn_image.save(msn_image_filename)
        pdf.drawImage(msn_image_filename, x2+310, y2-75, width=barcode_width, height=barcode_height)
        pdf.drawString(x2+260, y2-50, f"MSN:")


        ean1 = Code128(ean, writer=ImageWriter())
        ean_image = ean1.render(writer_options={'module_width': 4, 'module_height': 80, "font_size": 20*5, "text_distance": 40, "quite_zone": 10})
        ean_image_filename = (f"./bufferDEL/ean_barcode.png")
        ean_image.save(ean_image_filename)
        pdf.drawImage(ean_image_filename, x2+310, y2-125, width=barcode_width, height=barcode_height)
        pdf.drawString(x2+260, y2-100, f"EAN:")



        pdf.setFont("Helvetica-Bold", font_size-1)
        
        ## RSN placed on the left side of the page. 

        # Insert barcode image into PDF
        rsn = Code128(val1[i], writer=ImageWriter())
        rsn_image = rsn.render(writer_options={'module_width': 2.8, 'module_height': 80, "font_size": 20*5  , "text_distance": 40, "quite_zone": 10})
        rsn_image_filename = (f"./bufferDEL/rsn_barcode_{i}.png")
        rsn_image.save(rsn_image_filename)
        pdf.drawImage(rsn_image_filename, x, y-18, width=barcode_width, height=barcode_height+12)
        pdf.drawString(x-35, y+25, f"RSN:")
        pdf.drawString(x-30, y+15, f"{no}")


        macid = Code128(val2[i], writer=ImageWriter())
        macid_image = macid.render(writer_options={'module_width': 2.8, 'module_height': 80, "font_size": 20*5, "text_distance": 40, "quite_zone": 10})
        macid_image_filename = (f"./bufferDEL/macid_barcode{i}.png")
        macid_image.save(macid_image_filename)
        pdf.drawImage(macid_image_filename, x1, y1-18, width=barcode_width, height=barcode_height+12)   
        pdf.drawString(x2+311, y2-130, f"MAC ID: ")
        
        







        # Remove the barcode image file
        os.remove(rsn_image_filename)
        os.remove(macid_image_filename)

        # Increment the barcode count
        barcode_count += 1

        # Check if a new page is needed
        if barcode_count == barcode_per_page:
            # Reset parameters for new page
            page_count += 1
            barcode_count = 0

            # Show the current page
            pdf.showPage()
        no = no+1
        i = i + 1 
    # Save the PDF
    print("File save in the script folder, look for cartonStickers.pdf")
    pdf.save()


def extract_data_below_values(sheet, search_value):
    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
        for cell in row:
            normalized_cell_value = str(cell.value).lower().replace(' ', '').replace('.', '').replace('\xa0', '')

            if search_value.lower() == normalized_cell_value:
                data_tuples = []
                for i in range(cell.row + 1, sheet.max_row + 1):
                    next_cell_value = sheet.cell(i, cell.column).value
                    if next_cell_value is None or "box" in str(next_cell_value).lower():
                        break
                    data_tuples.append((search_value, next_cell_value))
                return data_tuples

    return f"Values not found for {search_value}"



# Load data from Excel file
workbook = openpyxl.load_workbook('MakeMe.xlsx')
sheet = workbook.active

#collecting data fomr excel file and storing it in an array
# val1 = []
# val2 = []

# row_num = 1
# while True:
#     cell1 = sheet.cell(row=row_num, column=1).value
#     cell2 = sheet.cell(row=row_num, column=2).value

#     if cell1 is None and cell2 is None:
#         break
    
#     while cell1 is not None and cell2 is not None:
#         val1.append(cell1)
#         val2.append(cell2)
#         break

#     row_num += 1


search_values = "box1"
print(search_values)
extracted_data = extract_data_below_values(sheet, search_values)



if isinstance(extracted_data, pd.DataFrame):
    print("Data found:")
    print(extracted_data)

    # Print the nth value, for example, the 3rd value (n=3)
    n = 3
    if n <= len(extracted_data):
        nth_value = extracted_data.iloc[n-1].item()  # n-1 because of zero-based indexing
        print(f"The {n}th value is: {nth_value}")
    else:
        print(f"There are less than {n} values in the extracted data.")
else:
    print(extracted_data)

#boxStickers(sheet, val1, val2)


